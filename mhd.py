import sys
import os
import openpyxl
from PyQt6.QtWidgets import QApplication, QMainWindow, QComboBox, QPushButton, QLabel, QVBoxLayout, QScrollArea, QWidget
from PyQt6.QtCore import QTimer, QTime, Qt
from PyQt6 import uic
import logging
import re


class TimetableApp(QMainWindow):
    def __init__(self):
        super().__init__()

        # Nastavení loggeru
        logging.basicConfig(
            level=logging.DEBUG,  # Nastavte požadovanou úroveň logování (DEBUG, INFO, WARNING, ERROR)
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[logging.StreamHandler()]  # Výstup logování do konzole
        )
        logging.info("Aplikace spuštěna")

        # Načtení UI
        if hasattr(sys, '_MEIPASS'):
            ui_file_path = os.path.join(sys._MEIPASS, 'mhd.ui')
        else:
            ui_file_path = os.path.join(os.path.dirname(__file__), 'mhd.ui')
        uic.loadUi(ui_file_path, self)

        # Přístup k widgetům
        self.combo_box = self.findChild(QComboBox, "comboBox")
        self.scroll_area = self.findChild(QScrollArea, "scrollArea")
        self.scroll_area_2 = self.findChild(QScrollArea, "scrollArea_2")
        self.label = self.findChild(QLabel, "label")

        # Nastavení dynamických cest
        self.app_dir = os.path.dirname(os.path.abspath(__file__))
        self.timetable_dir = os.path.join(self.app_dir, "Jízdní řády")

        # Inicializace layoutů
        self.scroll_area_widget = QWidget()
        self.scroll_area.setWidget(self.scroll_area_widget)
        self.scroll_area_layout = QVBoxLayout(self.scroll_area_widget)

        self.scroll_area_2_widget = QWidget()
        self.scroll_area_2.setWidget(self.scroll_area_2_widget)
        self.scroll_area_2_layout = QVBoxLayout(self.scroll_area_2_widget)

        # Načtení seznamu jízdních řádů
        self.update_combobox()

        # Připojení signálu pro comboBox
        self.combo_box.currentTextChanged.connect(self.load_selected_timetable)

        # Časovač pro aktualizaci odpočtu
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_countdown)
        self.timer.start(1000)
        
        #self.load_selected_timetable(self.combo_box.currentText())

        # Inicializace proměnných
        self.current_times = []
        self.current_stop = None
        self.timetable = {}

    def update_combobox(self):
        """Načte seznam jízdních řádů do comboBoxu a přidá výchozí položku."""
        self.combo_box.clear()
        if not os.path.exists(self.timetable_dir):
            os.makedirs(self.timetable_dir)  # Vytvoří složku, pokud neexistuje
    
        # Načteme pouze platné XLSX soubory a ignorujeme dočasné (~$) soubory
        xlsx_files = [
            f[:-5] for f in os.listdir(self.timetable_dir)
            if f.endswith(".xlsx") and not f.startswith("~$")
        ]
    
        # Seřadíme soubory podle čísel na začátku názvu
        def extract_number(name):
            match = re.match(r"(\d+)", name)  # Najde číslo na začátku názvu
            return int(match.group(1)) if match else float("inf")
    
        xlsx_files.sort(key=extract_number)
    
        # Přidáme výchozí statickou položku
        self.combo_box.addItem("Vyberte si linku")
        self.combo_box.addItems(xlsx_files)
    
        # Nastavení výchozí hodnoty
        self.combo_box.setCurrentIndex(0)

    def load_selected_timetable(self, selected_timetable):
        """Načte vybraný jízdní řád z XLSX."""
        if not selected_timetable:
            return
        xlsx_path = os.path.join(self.timetable_dir, f"{selected_timetable}.xlsx")
        if os.path.exists(xlsx_path):
            self.timetable = self.parse_xlsx_timetable(xlsx_path)
            self.generate_stop_buttons()

    def parse_xlsx_timetable(self, xlsx_path):
        """Načte jízdní řád z XLSX a vrátí slovník zastávek s časy."""
        logging.info(f"Načítám XLSX: {xlsx_path}")
        timetable = {}
        workbook = openpyxl.load_workbook(xlsx_path)
        for sheet in workbook.sheetnames:
            logging.info(f"Zpracovávám list: {sheet}")
            sheet_data = workbook[sheet]
            stop_name = sheet.strip()
            timetable[stop_name] = []
            for row_index, row in enumerate(sheet_data.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
                hour = row[0]
                minutes = row[1]
                if hour is not None and minutes:
                    try:
                        minute_list = [minute.strip() for minute in str(minutes).split(",")]
                        for minute in minute_list:
                            timetable[stop_name].append(f"{int(hour):02}:{minute}")
                    except Exception as e:
                        logging.error(f"Chyba při zpracování řádku {row_index} v listu {sheet}: {e}")
                else:
                    logging.warning(f"Prázdná nebo neplatná hodnota na řádku {row_index} v listu {sheet}")
        return timetable

    def generate_stop_buttons(self):
        """Generuje tlačítka pro zastávky na základě jízdního řádu a nastavuje stylesheet."""
        # Vymazání starých tlačítek
        for i in reversed(range(self.scroll_area_layout.count())):
            self.scroll_area_layout.itemAt(i).widget().deleteLater()
    
        # Definování stylesheetu pro tlačítka
        button_stylesheet = """
        QPushButton {
            border-width: 0px;
            border-radius: 0px;
            border-color: #4d4c4c;
            font-size: 20px;
        }
    
        QPushButton:hover {
            background-color: rgb(182, 182, 182);
            color: rgb(0, 0, 0);
        }
    
        QPushButton:pressed {
            border-style: inset;
            background: qradialgradient(
                cx: 0.4, cy: -0.1, fx: 0.4, fy: -0.1,
                radius: 1.35, stop: 0 #fff, stop: 1 #ddd
            );
        }
        """
    
        # Generování tlačítek pouze s názvem zastávky
        for stop in self.timetable.keys():
            stop_cleaned = stop.strip().replace("\xa0", " ")  # Odstraníme nadbytečné mezery a nahradíme non-breaking space
            button = QPushButton(stop_cleaned)
            button.setStyleSheet(button_stylesheet)  # Použití stylu na tlačítko
            button.clicked.connect(lambda checked, s=stop_cleaned: self.show_times(s))
            self.scroll_area_layout.addWidget(button)
            logging.info(f"Tlačítko vytvořeno: {stop_cleaned}")

    def show_times(self, stop):
        """Zobrazuje časy příjezdu pro vybranou zastávku."""
        stop = stop.strip().replace("\xa0", " ")  # Odstranění mezer a non-breaking space
        self.current_stop = stop
        if stop not in self.timetable:
            logging.error(f"Zastávka {stop} nebyla nalezena ve slovníku jízdního řádu")
            return
    
        current_time = QTime.currentTime()
    
        # Filtruje časy příjezdu od aktuálního času
        self.current_times = [
            time for time in self.timetable.get(stop, [])
            if QTime.fromString(time, "HH:mm") >= current_time
        ]
    
        # Pokud nejsou časy k dispozici, zobrazí všechny
        if not self.current_times:
            self.current_times = self.timetable.get(stop, [])
    
        # Vymazání starých časů z widgetu
        for i in reversed(range(self.scroll_area_2_layout.count())):
            self.scroll_area_2_layout.itemAt(i).widget().deleteLater()
    
        # Přidání nových časů jako QLabel
        for time in self.current_times:
            label = QLabel(time)
            self.scroll_area_2_layout.addWidget(label)
    
        self.scroll_area_2_layout.update()
        self.update_countdown()
        logging.info(f"Zobrazeny časy pro zastávku: {stop}")

    def update_countdown(self):
        """Aktualizuje odpočet do nejbližšího času příjezdu a zobrazí aktuální časy."""
        if not self.current_stop:
            self.label.setText("Odpočet: -")
            return
    
        # Znovu načítáme časy příjezdu od aktuálního času
        current_time = QTime.currentTime()
        self.current_times = [
            time for time in self.timetable.get(self.current_stop, [])
            if QTime.fromString(time, "HH:mm") >= current_time
        ]
    
        # Pokud nejsou časy k dispozici, zobrazí všechny
        if not self.current_times:
            self.current_times = self.timetable.get(self.current_stop, [])
    
        # Vymazání starých časů z widgetu
        for i in reversed(range(self.scroll_area_2_layout.count())):
            self.scroll_area_2_layout.itemAt(i).widget().deleteLater()
    
        # Přidání nových časů jako QLabel
        for time in self.current_times:
            label = QLabel(time)
            self.scroll_area_2_layout.addWidget(label)
    
        self.scroll_area_2_layout.update()
    
        # Aktualizace odpočtu
        if self.current_times:
            next_time = QTime.fromString(self.current_times[0], "HH:mm")
            minutes_to_next = current_time.secsTo(next_time) // 60
    
            if minutes_to_next >= 0:
                self.label.setText(f"Příjezd za: {minutes_to_next} min")
            else:
                self.label.setText("Odpočet: -")
        else:
            self.label.setText("Odpočet: -")



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TimetableApp()
    window.show()
    sys.exit(app.exec())
